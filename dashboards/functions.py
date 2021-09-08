# Utilities
from dashboards.models import Vehiculo
from datetime import date, datetime, timedelta
from heapq import nlargest
from openpyxl import load_workbook

def excel():
    
    """FILE_PATH = "C://Users/pc/Downloads/economicos - llantas.xlsx"
    SHEET = "Hoja1"
    workbook = load_workbook(FILE_PATH, read_only=True)
    sheet = workbook[SHEET]

    for row in sheet.iter_rows(min_row=2):
        numero_economico = row[0].value
        modelo = row[1].value
        marca = row[2].value
        compañia = row[3].value
        clase = row[4].value
        ubicacion = row[5].value
        aplicacion = row[6].value
        presion_establecida = row[7].value
        Vehiculo.objects.create(numero_economico=numero_economico, 
                                modelo=modelo, 
                                marca=marca, 
                                compañia=compañia, 
                                clase=clase,
                                ubicacion=ubicacion,
                                aplicacion=aplicacion,
                                presion_establecida=presion_establecida
                            )
    """

    """while Vehiculo.objects.count():
        ids = Vehiculo.objects.values_list('pk', flat=True)[:100]
        Vehiculo.objects.filter(pk__in = ids).delete()
    """

def clases_mas_frecuentes():
    vehiculo = Vehiculo.objects.all()
    clases = {}
    for v in vehiculo:
        clase = v.clase.capitalize()
        if clase in clases:
            clases[clase] += 1
        else:
            clases[clase] = 1
    clases.pop("")
    clases_mayores = nlargest(8, clases, key=clases.get)
    for c in clases.copy():
        if c in clases_mayores:
            pass
        else:
            clases.pop(c)
    return clases


def contar_dias(fecha):
    fecha_date = datetime.strptime(fecha, "%Y-%m-%d").date()
    hoy = date.today()
    return hoy - fecha_date



def contar_entrada_correcta(vehiculo_fecha):
    entrada_correcta_contar = 0
    for vehiculo_entrada_correcta in vehiculo_fecha:
        presion_encontrada = vehiculo_entrada_correcta.presion_de_entrada
        presion_establecida = vehiculo_entrada_correcta.presion_establecida
        entrada_correcta = presion_encontrada/presion_establecida

        if entrada_correcta >= 0.9:
            entrada_correcta_contar += 1

    return entrada_correcta_contar



def convertir_fecha(fecha):
    partes_fecha = fecha.split("-")
    return f"{partes_fecha[2]}/{partes_fecha[1]}/{partes_fecha[0][2:4]}"

def convertir_rango(fecha, fecha2):
    partes_fecha = fecha.split("-")
    partes_fecha2 = fecha2.split("-")
    return f"{partes_fecha[2]}/{partes_fecha[1]}/{partes_fecha[0][2:4]} - {partes_fecha2[2]}/{partes_fecha2[1]}/{partes_fecha2[0][2:4]}"
    


def inflado_promedio(vehiculo):
    tiempo_promedio = 0
    promedio_contar = 0
    for tiempo in vehiculo:
        tiempo_promedio += tiempo.tiempo_de_inflado
        promedio_contar +=1
    try:
        return round(tiempo_promedio/promedio_contar, 2)
    except:
        return None



def porcentaje(divisor, dividendo):
    try:
        return int((divisor/dividendo)*100)
    except:
        return "Nada"

def radar_min():
    clases = clases_mas_frecuentes()
    clases_min = min(clases.keys(), key=lambda k: clases[k])
    return clases[clases_min]

def radar_max():
    clases = clases_mas_frecuentes()
    clases_max = max(clases.keys(), key=lambda k: clases[k])
    return clases[clases_max]

def radar_min_resta(min, max):
    resta = max - min
    resta = resta // 2.3
    return min - resta